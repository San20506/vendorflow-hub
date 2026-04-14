#!/usr/bin/env python3
"""
Import client marketplace data into Supabase platform_orders + platform_settlements tables.
Run: python3 scripts/import_client_data.py

Requires: pip install supabase openpyxl
"""

import os
import sys
import csv
import json
import glob
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from supabase import create_client, Client
except ImportError:
    print("Installing dependencies...")
    os.system(f"{sys.executable} -m pip install supabase openpyxl -q")
    import openpyxl
    from supabase import create_client, Client

# ─── Config ─────────────────────────────────────────────────────────────────
DATA_DIR = Path(__file__).parent.parent / "Data" / "extracted"
SUPABASE_URL = os.environ.get("VITE_SUPABASE_URL", "https://omtjwinoxbrzeqvdckyg.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("VITE_SUPABASE_PUBLISHABLE_KEY", "")

# Load from .env if not set
if not SUPABASE_KEY:
    env_path = Path(__file__).parent.parent / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.startswith("SUPABASE_SERVICE_ROLE_KEY="):
                SUPABASE_KEY = line.split("=", 1)[1].strip().strip('"')
                break
            if not SUPABASE_KEY and line.startswith("VITE_SUPABASE_PUBLISHABLE_KEY="):
                SUPABASE_KEY = line.split("=", 1)[1].strip().strip('"')

VENDOR_ID = os.environ.get("VENDOR_ID", "")  # will auto-fetch first vendor if empty

print(f"Supabase URL: {SUPABASE_URL}")
print(f"Key: {SUPABASE_KEY[:30]}...")

# ─── Helpers ─────────────────────────────────────────────────────────────────

def clean_str(v):
    if v is None: return None
    s = str(v).strip().replace('"""', '').strip('"')
    return s if s and s != '-' and s != 'None' else None

def clean_num(v):
    if v is None: return None
    try: return float(str(v).replace(',', ''))
    except: return None

def clean_date(v):
    if v is None: return None
    s = str(v).strip()
    if not s or s == '-': return None
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
        try:
            return datetime.strptime(s[:len(fmt)], fmt).isoformat()
        except: pass
    return s  # return as-is if can't parse

def xlsx_rows(path, sheet_name=None, header_row=0, data_start=None):
    """Yield (headers, row_dict) tuples from xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows: return
    headers = [str(h) if h is not None else '' for h in rows[header_row]]
    start = data_start if data_start is not None else header_row + 1
    for row in rows[start:]:
        if all(v is None for v in row): continue
        yield headers, list(row)

def batch_upsert(supabase: Client, table: str, records: list, batch_size=200):
    inserted = 0
    skipped = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        try:
            res = supabase.table(table).upsert(batch, on_conflict="vendor_id,platform,platform_order_id,platform_order_item_id").execute()
            inserted += len(batch)
        except Exception as e:
            err = str(e)
            if "duplicate" in err.lower() or "unique" in err.lower():
                skipped += len(batch)
            else:
                print(f"  ERROR in batch {i//batch_size}: {err[:120]}")
                skipped += len(batch)
    return inserted, skipped

def batch_upsert_settlements(supabase: Client, records: list, batch_size=200):
    inserted = 0
    skipped = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        try:
            res = supabase.table("platform_settlements").upsert(batch, on_conflict="vendor_id,platform,settlement_ref,platform_order_id").execute()
            inserted += len(batch)
        except Exception as e:
            err = str(e)
            if "duplicate" in err.lower() or "unique" in err.lower():
                skipped += len(batch)
            else:
                print(f"  ERROR in settlement batch: {err[:120]}")
                skipped += len(batch)
    return inserted, skipped

# ─── Flipkart Orders ─────────────────────────────────────────────────────────

def parse_flipkart_orders(path, vendor_id):
    records = []
    try:
        for headers, row in xlsx_rows(path, sheet_name="Orders", header_row=0):
            def g(k): return row[headers.index(k)] if k in headers else None
            oid = clean_str(g('order_id'))
            if not oid: continue
            sku = clean_str(g('sku'))
            if sku: sku = re.sub(r'^SKU:', '', sku)
            records.append({
                "vendor_id": vendor_id,
                "platform": "flipkart",
                "platform_order_id": oid,
                "platform_order_item_id": clean_str(g('order_item_id')) or oid,
                "order_date": clean_date(g('order_date')),
                "sku": sku,
                "product_name": clean_str(g('product_title')),
                "quantity": int(g('quantity') or 1),
                "status": clean_str(g('order_item_status')),
                "return_reason": clean_str(g('return_reason')),
                "fulfillment_type": clean_str(g('fulfilment_type')),
                "raw_data": {h: str(v) for h, v in zip(headers, row) if v is not None},
            })
    except Exception as e:
        print(f"  Parse error {path.name}: {e}")
    return records

# ─── Flipkart Settlements ────────────────────────────────────────────────────

def parse_flipkart_settlements(path, vendor_id):
    records = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Orders" not in wb.sheetnames:
            wb.close(); return records
        ws = wb["Orders"]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(all_rows) < 3: return records
        row1 = [str(v) if v else '' for v in all_rows[0]]
        row2 = [str(v).strip() if v else '' for v in all_rows[1]]
        for row in all_rows[3:]:
            if all(v is None for v in row): continue
            def g(key):
                idx = next((i for i, h in enumerate(row2) if key.lower() in h.lower()), -1)
                return row[idx] if idx >= 0 else None
            oid = clean_str(g('Order ID'))
            if not oid or oid == '-': continue
            net = clean_num(g('Bank Settlement Value'))
            sale = clean_num(g('Sale Amount'))
            comm = clean_num(g('Commission'))
            ship = clean_num(g('Shipping Fee'))
            tcs = clean_num(g('TCS'))
            tds = clean_num(g('TDS'))
            ref = clean_str(g('NEFT ID'))
            pdate = clean_date(g('Payment Date'))
            records.append({
                "vendor_id": vendor_id,
                "platform": "flipkart",
                "settlement_ref": ref or f"FK-{oid}",
                "platform_order_id": oid,
                "payment_date": pdate,
                "sale_amount": abs(sale) if sale else None,
                "commission": abs(comm) if comm else None,
                "shipping_fee": abs(ship) if ship else None,
                "tcs": abs(tcs) if tcs else None,
                "tds": abs(tds) if tds else None,
                "net_settlement": net,
                "raw_data": {row2[i]: str(v) for i, v in enumerate(row) if v is not None and row2[i]},
            })
    except Exception as e:
        print(f"  Settlement parse error {path.name}: {e}")
    return records

# ─── Meesho Orders + Settlements ─────────────────────────────────────────────

def parse_meesho_payments(path, vendor_id):
    orders, settlements = [], []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Order Payments" not in wb.sheetnames:
            wb.close(); return orders, settlements
        ws = wb["Order Payments"]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(all_rows) < 4: return orders, settlements
        headers = [str(v).strip() if v else '' for v in all_rows[1]]
        def g(row, key):
            idx = next((i for i, h in enumerate(headers) if h.startswith(key)), -1)
            return row[idx] if idx >= 0 else None
        for row in all_rows[3:]:
            if all(v is None for v in row): continue
            sub = clean_str(g(row, 'Sub Order No'))
            if not sub or sub == '-': continue
            qty = int(g(row, 'Quantity') or 1)
            listing = clean_num(g(row, 'Listing Price'))
            total_sale = clean_num(g(row, 'Total Sale Amount'))
            final = clean_num(g(row, 'Final Settlement Amount'))
            orders.append({
                "vendor_id": vendor_id,
                "platform": "meesho",
                "platform_order_id": sub,
                "platform_order_item_id": sub,
                "order_date": clean_date(g(row, 'Order Date')),
                "sku": clean_str(g(row, 'Supplier SKU')),
                "product_name": clean_str(g(row, 'Product Name')),
                "quantity": qty,
                "mrp": listing,
                "sale_amount": total_sale or listing,
                "status": clean_str(g(row, 'Live Order Status')),
                "raw_data": {headers[i]: str(v) for i, v in enumerate(row) if v is not None and headers[i]},
            })
            txn = clean_str(g(row, 'Transaction ID'))
            comm = clean_num(g(row, 'Meesho Commission'))
            ship = clean_num(g(row, 'Shipping Charge'))
            rship = clean_num(g(row, 'Return Shipping Charge'))
            tcs = clean_num(g(row, 'TCS'))
            tds = clean_num(g(row, 'TDS'))
            fixed = clean_num(g(row, 'Fixed Fee'))
            settlements.append({
                "vendor_id": vendor_id,
                "platform": "meesho",
                "settlement_ref": txn or f"ME-{sub}",
                "platform_order_id": sub,
                "payment_date": clean_date(g(row, 'Payment Date')),
                "sale_amount": total_sale,
                "commission": abs(comm) if comm else None,
                "shipping_fee": (abs(ship or 0) + abs(rship or 0)) or None,
                "tcs": abs(tcs) if tcs else None,
                "tds": abs(tds) if tds else None,
                "other_deductions": abs(fixed) if fixed else None,
                "net_settlement": final,
                "raw_data": {headers[i]: str(v) for i, v in enumerate(row) if v is not None and headers[i]},
            })
    except Exception as e:
        print(f"  Meesho parse error {path.name}: {e}")
    return orders, settlements

# ─── Firstcry Orders ─────────────────────────────────────────────────────────

def parse_firstcry_orders(path, vendor_id):
    records = []
    try:
        for headers, row in xlsx_rows(path, header_row=0):
            def g(k): return row[headers.index(k)] if k in headers else None
            poid = clean_str(g('POID'))
            if not poid: continue
            records.append({
                "vendor_id": vendor_id,
                "platform": "firstcry",
                "platform_order_id": poid,
                "platform_order_item_id": poid,
                "order_date": clean_date(g('OrderDate')),
                "sku": clean_str(g('VendorStyleCode')) or clean_str(g('ProductID')),
                "product_name": clean_str(g('Brand Name')),
                "quantity": int(g('Quantity') or 1),
                "mrp": clean_num(g('MRP')),
                "sale_amount": clean_num(g('MRP Sales')) or clean_num(g('MRP')),
                "category": clean_str(g('categoryname')),
                "subcategory": clean_str(g('subcategoryname')),
                "raw_data": {h: str(v) for h, v in zip(headers, row) if v is not None},
            })
    except Exception as e:
        print(f"  Firstcry orders parse error: {e}")
    return records

# ─── Firstcry Settlements ─────────────────────────────────────────────────────

def parse_firstcry_reconciliation(path, vendor_id):
    records = []
    try:
        for headers, row in xlsx_rows(path, header_row=0):
            def g(k): return row[headers.index(k)] if k in headers else None
            ref = clean_str(g('FC Ref. no.'))
            if not ref: continue
            records.append({
                "vendor_id": vendor_id,
                "platform": "firstcry",
                "settlement_ref": clean_str(g('Payment advice no')) or ref,
                "platform_order_id": clean_str(g('Order Ids')),
                "payment_date": clean_date(g('Delivery date')),
                "sale_amount": clean_num(g('Gross Amount')),
                "tcs": (clean_num(g('CGST Amount')) or 0) + (clean_num(g('SGST Amount')) or 0) or None,
                "other_deductions": (clean_num(g('SR Total Amount')) or 0) + (clean_num(g('RTO Total Amount')) or 0) or None,
                "net_settlement": clean_num(g('Total')),
                "raw_data": {h: str(v) for h, v in zip(headers, row) if v is not None},
            })
    except Exception as e:
        print(f"  Firstcry recon parse error: {e}")
    return records

# ─── Amazon ──────────────────────────────────────────────────────────────────

def parse_amazon_metrics(path, vendor_id):
    records = []
    try:
        for headers, row in xlsx_rows(path, header_row=0):
            def g(k):
                idx = next((i for i, h in enumerate(headers) if h.strip() == k), -1)
                return row[idx] if idx >= 0 else None
            asin = clean_str(g('ASIN'))
            if not asin: continue
            units = clean_num(g('Units sold')) or 0
            if units <= 0: continue  # skip zero-sales rows
            start = clean_str(g('Start date'))
            records.append({
                "vendor_id": vendor_id,
                "platform": "amazon",
                "platform_order_id": f"{asin}-{start}",
                "platform_order_item_id": f"{asin}-{start}",
                "order_date": clean_date(start),
                "sku": clean_str(g('MSKU')) or asin,
                "product_name": None,
                "quantity": int(units),
                "mrp": clean_num(g('Average sales price')),
                "sale_amount": clean_num(g('Net sales')) or clean_num(g('Sales')),
                "status": "DELIVERED",
                "raw_data": {headers[i].strip(): str(v) for i, v in enumerate(row) if v is not None and headers[i]},
            })
    except Exception as e:
        print(f"  Amazon metrics parse error {path.name}: {e}")
    return records

def parse_amazon_returns(path, vendor_id):
    records = []
    try:
        with open(path, encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter='\t')
            for row in reader:
                oid = row.get('Order ID', '').strip()
                if not oid: continue
                records.append({
                    "vendor_id": vendor_id,
                    "platform": "amazon",
                    "platform_order_id": oid,
                    "platform_order_item_id": oid,
                    "order_date": clean_date(row.get('Order date') or row.get('Return request date')),
                    "sku": row.get(' Merchant SKU', row.get('Merchant SKU', '')).strip(),
                    "product_name": row.get('Item Name', '').strip()[:255],
                    "quantity": int(row.get('Return quantity', 1) or 1),
                    "status": "RETURNED",
                    "return_reason": row.get('Return reason', '').strip(),
                    "raw_data": dict(row),
                })
    except Exception as e:
        print(f"  Amazon returns parse error {path.name}: {e}")
    return records

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    global VENDOR_ID

    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Fetch vendor_id if not set
    if not VENDOR_ID:
        res = supabase.table("vendors").select("vendor_id,business_name").limit(1).execute()
        if not res.data:
            print("ERROR: No vendors found. Seed test accounts first via /admin-tools.")
            sys.exit(1)
        VENDOR_ID = res.data[0]["vendor_id"]
        print(f"Using vendor: {res.data[0].get('business_name', VENDOR_ID)} ({VENDOR_ID})")

    total_orders = 0
    total_settlements = 0

    # ── Flipkart ──────────────────────────────────────────────
    print("\n[1/4] Flipkart")
    fk_dir = DATA_DIR / "flipkart" / "Flipkart Sept_Oct_Nov_Dec_2025"
    for f in sorted(fk_dir.glob("*_Orders.xlsx")):
        print(f"  Orders: {f.name}")
        recs = parse_flipkart_orders(f, VENDOR_ID)
        if recs:
            ins, skip = batch_upsert(supabase, "platform_orders", recs)
            print(f"    → {ins} inserted, {skip} skipped")
            total_orders += ins

    for f in sorted(fk_dir.glob("*_Settled Transactions.xlsx")):
        print(f"  Settlements: {f.name}")
        recs = parse_flipkart_settlements(f, VENDOR_ID)
        if recs:
            ins, skip = batch_upsert_settlements(supabase, recs)
            print(f"    → {ins} inserted, {skip} skipped")
            total_settlements += ins

    # ── Meesho ────────────────────────────────────────────────
    print("\n[2/4] Meesho")
    me_dir = DATA_DIR / "meesho"
    for f in sorted(me_dir.glob("meesho_PREVIOUS_PAYMENT_*/*.xlsx")):
        print(f"  Payment: {f.parent.name}")
        orders, settlements = parse_meesho_payments(f, VENDOR_ID)
        if orders:
            ins, skip = batch_upsert(supabase, "platform_orders", orders)
            print(f"    → Orders: {ins} inserted, {skip} skipped")
            total_orders += ins
        if settlements:
            ins, skip = batch_upsert_settlements(supabase, settlements)
            print(f"    → Settlements: {ins} inserted, {skip} skipped")
            total_settlements += ins

    # Outstanding payments
    for f in sorted(me_dir.glob("meesho_OUTSTANDING_PAYMENT_*/*.xlsx")):
        print(f"  Outstanding: {f.parent.name}")
        orders, settlements = parse_meesho_payments(f, VENDOR_ID)
        if orders:
            ins, skip = batch_upsert(supabase, "platform_orders", orders)
            print(f"    → Orders: {ins} inserted, {skip} skipped")
            total_orders += ins

    # ── Firstcry ──────────────────────────────────────────────
    print("\n[3/4] Firstcry")
    fc_dir = DATA_DIR / "firstcry" / "Firstcry"
    for f in fc_dir.glob("dashboardsale*.xlsx"):
        print(f"  Orders: {f.name}")
        recs = parse_firstcry_orders(f, VENDOR_ID)
        if recs:
            ins, skip = batch_upsert(supabase, "platform_orders", recs)
            print(f"    → {ins} inserted, {skip} skipped")
            total_orders += ins

    for f in sorted(fc_dir.glob("VendorReconciliation*.xlsx")):
        print(f"  Reconciliation: {f.name}")
        recs = parse_firstcry_reconciliation(f, VENDOR_ID)
        if recs:
            ins, skip = batch_upsert_settlements(supabase, recs)
            print(f"    → {ins} inserted, {skip} skipped")
            total_settlements += ins

    # ── Amazon ────────────────────────────────────────────────
    print("\n[4/4] Amazon")
    az_dir = DATA_DIR / "amazon" / "Amazon"
    for f in sorted(az_dir.glob("metric-data*.xlsx")):
        print(f"  Metrics: {f.name}")
        recs = parse_amazon_metrics(f, VENDOR_ID)
        if recs:
            ins, skip = batch_upsert(supabase, "platform_orders", recs)
            print(f"    → {ins} inserted, {skip} skipped")
            total_orders += ins

    for f in sorted(az_dir.glob("report-*.tsv")):
        print(f"  Returns: {f.name}")
        recs = parse_amazon_returns(f, VENDOR_ID)
        if recs:
            ins, skip = batch_upsert(supabase, "platform_orders", recs)
            print(f"    → {ins} inserted, {skip} skipped")
            total_orders += ins

    print(f"\n✓ Done. Total: {total_orders} orders, {total_settlements} settlements imported.")

if __name__ == "__main__":
    main()
